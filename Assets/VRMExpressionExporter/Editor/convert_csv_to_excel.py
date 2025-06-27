#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VRM Expression CSV to Excel Converter
CSVファイルから画像が埋め込まれたExcelファイルを生成します。
"""

import os
import sys
import csv
from collections import defaultdict
import argparse
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is not installed. Please run: pip install openpyxl")
    sys.exit(1)

try:
    from PIL import Image
except ImportError:
    print("Error: Pillow is not installed. Please run: pip install Pillow")
    sys.exit(1)


def read_csv_file(csv_path):
    """CSVファイルを読み込んでデータを返す"""
    data = []
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            data.append(row)
    return data


def group_data_by_character(data):
    """キャラクターごとにデータをグループ化"""
    grouped = defaultdict(list)
    for row in data:
        char_name = row.get('オブジェクト名', 'Unknown')
        grouped[char_name].append(row)
    return grouped


def resize_image(image_path, max_width=150, max_height=150):
    """画像をリサイズしてバイトデータとして返す（高解像度を維持）"""
    try:
        from io import BytesIO
        img = Image.open(image_path)
        
        # RGBA形式に変換（透過をサポート）
        if img.mode != 'RGBA':
            img = img.convert('RGBA')
        
        # アスペクト比を保持してリサイズ（高品質なリサンプリング）
        img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
        
        # バイトストリームに保存（高品質設定）
        img_buffer = BytesIO()
        # PNG形式で保存（最適化して品質を維持）
        img.save(img_buffer, format='PNG', optimize=True)
        img_buffer.seek(0)
        
        return img_buffer
    except Exception as e:
        print(f"Warning: Failed to load image {image_path}: {e}")
        return None


def create_excel_file(csv_path, output_path=None):
    """CSVファイルからExcelファイルを作成"""
    # 出力パスの設定
    if output_path is None:
        output_path = csv_path.replace('.csv', '_with_images.xlsx')
    
    # プロジェクトのルートディレクトリを取得
    project_root = Path(csv_path).parent.parent
    
    # CSVデータを読み込み
    print(f"Reading CSV file: {csv_path}")
    data = read_csv_file(csv_path)
    
    # Excelワークブックを作成
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除
    
    # キャラクターごとにデータをグループ化
    grouped_data = group_data_by_character(data)
    
    # 各キャラクターごとにシートを作成
    for char_name, char_data in grouped_data.items():
        print(f"Processing character: {char_name}")
        
        # シート名を作成（Excelのシート名制限に対応）
        sheet_name = char_name[:31] if len(char_name) > 31 else char_name
        sheet_name = sheet_name.replace('/', '_').replace('\\', '_')
        ws = wb.create_sheet(title=sheet_name)
        
        # ヘッダーを設定
        headers = ['表情名', '画像', 'ブレンドシェイプパス', 'ブレンドシェイプ名', '値(%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 表情ごとにデータをグループ化
        expression_groups = defaultdict(list)
        for row in char_data:
            expr_name = row.get('表情名', '')
            expression_groups[expr_name].append(row)
        
        # データ行を追加
        current_row = 2
        for expr_name, expr_rows in expression_groups.items():
            start_row = current_row
            
            # 表情名を記入
            ws.cell(row=current_row, column=1, value=expr_name)
            
            # 画像を挿入
            if expr_rows and '画像パス' in expr_rows[0]:
                image_path = os.path.join(project_root, expr_rows[0]['画像パス'])
                if os.path.exists(image_path):
                    try:
                        # 画像をリサイズ（高解像度を維持）
                        # 実際の画像は200x200で保持し、表示は小さくする
                        img_buffer = resize_image(image_path, max_width=200, max_height=200)
                        if img_buffer:
                            # Excelに画像を挿入
                            xl_img = XLImage(img_buffer)
                            
                            # 表示サイズを設定（実際の画像より小さく表示）
                            # これにより高解像度を維持しながらセル内に収める
                            xl_img.width = 80
                            xl_img.height = 80
                            
                            # セルの中央に配置するため、少しオフセットを設定
                            # B列のcurrent_row行に画像を配置
                            # アンカーをセルに設定（openpyxlの標準的な方法）
                            cell_address = f'B{current_row}'
                            
                            # 画像を追加（セルに紐付け）
                            ws.add_image(xl_img, cell_address)
                            
                            # 行の高さを画像に合わせて調整（余白込み）
                            ws.row_dimensions[current_row].height = 65
                            
                            # セルに値を設定して、画像の存在を示す
                            # （これにより画像がセルに関連付けられる）
                            ws.cell(row=current_row, column=2, value="")
                    except Exception as e:
                        print(f"Warning: Failed to insert image for {expr_name}: {e}")
                        ws.cell(row=current_row, column=2, value="画像読み込みエラー")
                else:
                    ws.cell(row=current_row, column=2, value="画像なし")
            
            # ブレンドシェイプ情報を記入
            for i, row in enumerate(expr_rows):
                if i > 0:  # 2行目以降
                    current_row += 1
                    ws.row_dimensions[current_row].height = 20
                
                ws.cell(row=current_row, column=3, value=row.get('ブレンドシェイプパス', ''))
                ws.cell(row=current_row, column=4, value=row.get('ブレンドシェイプ名', ''))
                ws.cell(row=current_row, column=5, value=row.get('値(%)', ''))
            
            # 表情名と画像をマージ
            if len(expr_rows) > 1:
                ws.merge_cells(f'A{start_row}:A{current_row}')
                ws.merge_cells(f'B{start_row}:B{current_row}')
                
                # マージしたセルの配置を調整
                ws[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'B{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
        
        # 列幅を調整（画像列を適切なサイズに）
        ws.column_dimensions['A'].width = 20  # 表情名
        ws.column_dimensions['B'].width = 14  # 画像（80x80の表示に合わせて調整）
        ws.column_dimensions['C'].width = 30  # ブレンドシェイプパス
        ws.column_dimensions['D'].width = 30  # ブレンドシェイプ名
        ws.column_dimensions['E'].width = 10  # 値
        
        # 罫線を追加
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:  # ヘッダー以外
                    cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Excelファイルを保存
    print(f"Saving Excel file: {output_path}")
    wb.save(output_path)
    print("Conversion completed successfully!")
    
    return output_path


def main():
    parser = argparse.ArgumentParser(description='Convert VRM Expression CSV to Excel with embedded images')
    parser.add_argument('csv_file', help='Path to the CSV file')
    parser.add_argument('-o', '--output', help='Output Excel file path (optional)')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.csv_file):
        print(f"Error: CSV file not found: {args.csv_file}")
        sys.exit(1)
    
    try:
        output_path = create_excel_file(args.csv_file, args.output)
        print(f"Excel file created: {output_path}")
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()